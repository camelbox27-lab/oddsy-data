"""
ODDSY - İddaa Günlük Maç Güncelleme Scripti
============================================
KULLANIM:
  1. gunlukmaclar.xlsx dosyasını bu klasöre koy (üzerine yaz)
  2. CALISTIR.bat'a çift tıkla
  3. JSON otomatik oluşturulur ve GitHub'a push edilir

GITHUB: camelbox27-lab/oddsy-data
JSON PATH: guncel_json/iddaa/gunlukmaclar.json
"""

import os
import sys
import json
import subprocess
from datetime import date

try:
    import openpyxl
except ImportError:
    print("[!] openpyxl bulunamadi, kuruluyor...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR   = os.path.dirname(os.path.dirname(SCRIPT_DIR))  # oddsy-data kök
EXCEL_FILE = os.path.join(SCRIPT_DIR, "gunlukmaclar.xlsx")
JSON_FILE  = os.path.join(SCRIPT_DIR, "gunlukmaclar.json")
GIT_PATH   = "guncel_json/iddaa/gunlukmaclar.json"

print("=" * 55)
print("  ODDSY - IDDAA GUNLUK MAC GUNCELLEME")
print(f"  {date.today().strftime('%d.%m.%Y')}")
print("=" * 55)

if not os.path.exists(EXCEL_FILE):
    print(f"\n[HATA] Excel bulunamadi: {EXCEL_FILE}")
    print("  gunlukmaclar.xlsx dosyasini bu klasore koy!")
    input("\nCikis icin Enter...\n")
    sys.exit(1)

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active
headers = [cell.value for cell in ws[1]]

rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row):
        continue
    entry = {}
    for key, val in zip(headers, row):
        entry[key] = None if val is None else str(val) if not isinstance(val, (int, float)) else val
    rows.append(entry)

with open(JSON_FILE, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print(f"\n[OK] {len(rows)} mac JSON'a yazildi -> gunlukmaclar.json")

os.chdir(REPO_DIR)

r = subprocess.run(["git", "add", GIT_PATH], capture_output=True, text=True)
if r.returncode != 0:
    print(f"[HATA] git add: {r.stderr}")
    input("\nCikis icin Enter...\n")
    sys.exit(1)

diff = subprocess.run(["git", "diff", "--cached", "--quiet"], capture_output=True)
if diff.returncode == 0:
    print("[BILGI] JSON degismedi, commit atiliyor.")
    input("\nCikis icin Enter...\n")
    sys.exit(0)

commit_msg = f"data: iddaa gunlukmaclar guncellendi ({date.today().strftime('%d.%m.%Y')}, {len(rows)} mac)"
r = subprocess.run(["git", "commit", "-m", commit_msg], capture_output=True, text=True)
if r.returncode != 0:
    print(f"[HATA] git commit: {r.stderr}")
    input("\nCikis icin Enter...\n")
    sys.exit(1)
print(f"[OK] Commit: {commit_msg}")

r = subprocess.run(["git", "push", "origin", "main"], capture_output=True, text=True)
if r.returncode != 0:
    print(f"[HATA] git push: {r.stderr}")
    input("\nCikis icin Enter...\n")
    sys.exit(1)

print("[OK] GitHub'a push edildi!")
print("=" * 55)
input("\nCikis icin Enter...\n")
